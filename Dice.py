import random
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
import math

class Dice:
    def __init__(self):
        self.dice = (1, 2, 3, 4, 5, 6)
        self.outcome = None
        self.games = 0
        self.theo = {2: 1 / 36, 3: 2 / 36, 4: 3 / 36, 5: 4 / 36, 6: 5 / 36, 7: 6 / 36, 8: 5 / 36, 9: 4 / 36, 10: 3 / 36,
                     11: 2 / 36, 12: 1 / 36} # theoretical probability of dice roll outcome 2 to 12.

    # roll dice and return dice1, dice1, result
    def roll(self):
        dice1 = random.choice(self.dice)
        dice2 = random.choice(self.dice)
        result = dice1 + dice2
        self.outcome = (dice1, dice2, result)
        return self.outcome

    # takes in the number of plays. Builds a list of outcomes
    def doutcomes(self, plays):
        outcomes = list()
        self.games = plays
        for x in range(1, plays + 1):
            outcomes.append(self.roll())
            # print(outcomes)
        return outcomes

    def exporttoexcel(self, resultslist):
        wb = Workbook()  # create new workbook
        worksheet1 = wb.active  # gets the default worksheet that was created when new workbook is created.
        count = 2
        worksheet1['A1'] = 'Dice 1'
        worksheet1['B1'] = 'Dice 2'
        worksheet1['C1'] = 'Outcome'
        for dice1, dice2, result in resultslist:
            worksheet1[f'A{count}'] = dice1
            worksheet1[f'B{count}'] = dice2
            worksheet1[f'C{count}'] = result
            count += 1
        return wb

    # use dictionary to keep tally for dice roll outcome
    def boutcomes(self, outcome):
        dic = dict()
        for dice1, dice2, result in outcome:
            dic[result] = dic.get(result, 0) + 1
        return dic

    # puts dictionary results back to list form to sort the outcomes.
    def extractd(self, temp):
        tmplist = list()
        for result, count in temp.items():
            tmplist.append((result, count))
        return tmplist

    # create chart from excel sheet data
    def chartme(self, wbook, histogram, name):
        rows = histogram
        sheet = wbook.active
        fsheet = wbook.create_sheet('temp')
        fsheet['A1'] = 'Outcome'
        fsheet['B1'] = 'Count'
        sheet['D1'] = 'Games Played'
        for row in rows:
            fsheet.append(row)
        sheet['D2'] = self.games

        # bar chart for actual probability
        data = Reference(fsheet, min_row=1, max_row=fsheet.max_row, min_col=2, max_col=2)
        titles = Reference(fsheet, min_row=2, max_row=fsheet.max_row, min_col=1, max_col=1)
        chart1 = BarChart()
        chart1.title = 'Dice Game'
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(titles)
        chart1.x_axis.title = 'Outcomes'
        chart1.y_axis.title = 'Frequency'
        # chart1.y_axis.scaling.max = 100
        # chart1.set_categories(titles)
        sheet.add_chart(chart1, 'g2')

        # populate fsheet with theo probability
        counter = 2
        for r, p in self.theo.items():
            fsheet[f'D{counter}'] = r
            fsheet[f'E{counter}'] = p
            counter += 1
        fsheet['E1'] = 'Probability'
            # bar chart for theo probability
        lchart = BarChart()
        lchart.title = 'Theo Probability for Dice Roll'
        theodata = Reference(fsheet, min_row=1, max_row=12, min_col=5, max_col=5)
        diceroll = Reference(fsheet, min_row=2, max_row=12, min_col=4, max_col=4)
        lchart.add_data(theodata,titles_from_data=True)
        lchart.set_categories(diceroll)
        sheet.add_chart(lchart, 'g20')
        wbook.save(f'{name}.xlsx')


user = input('How many games do you want to play? ')
fname = input('What file name will you use to save? ')
game1 = Dice()
outcome = game1.doutcomes(int(user))
newlist = game1.boutcomes(outcome)  # builds histogram
final = game1.extractd(newlist)  # builds list from histogram
sortedfinal = sorted(final)
workbook = game1.exporttoexcel(outcome)

# print(f'{outcome}\n')
# print(f'{final}\n')
game1.chartme(workbook, sortedfinal, fname)

# game2 = Dice()
# outcome2 = game2.doutcomes(1000)
# newlist2 = game2.boutcomes(outcome2)  # builds histogram
# final2 = game2.extractd(newlist2)  # builds list from histogram
# workbook2 = game2.exporttoexcel(outcome2)
# print(f'{outcome2}\n')
# print(f'{final2}\n')
# game2.chartme(workbook2, 'Outcomes', 'Frequency', 'Dice Game', final2, 'Game2')
