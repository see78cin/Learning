from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference


fname = input('Enter file name: ')
# strip out the punctuation at end of sentence and return a list words
def mylist(fname):
    l = list()
    file = open(fname)
    for line in file:
        line = line.rstrip()
        if len(line) < 1:
            continue
        words = line.split()
        for word in words:
            if '.' in word:
                l.append(word[:word.rfind('.')])
            elif '?' in word:
                l.append(word[:word.rfind('?')])
            elif '!' in word:
                l.append(word[:word.rfind('!')])
            elif ',' in word:
                l.append(word[:word.rfind(',')])
            else:
                l.append(word)
                print(l)
    return l

# build a dictionary. Takes in list
def chist(lis):
    dic = dict()
    for word in lis:
        dic[word] = dic.get(word, 0) + 1
    return dic


# takes a dictionary and return a list with k v swapped
def extractd(temp):
    tmplist = list()
    for k, v in temp.items():
        tmplist.append((v, k))
    return tmplist


def exporttoexcel(alist):
    wb = Workbook()  # create new workbook
    worksheet1 = wb.active  # creates a new worksheet.
    count = 2
    worksheet1['B1'] = 'Count'
    for k, v in alist[:5]:
        worksheet1[f'A{count}'] = v
        worksheet1[f'B{count}'] = k
        count += 1
    return wb


def chartme(wbook, x_title, y_title,chart_title):
    sheet = wbook.active
    data = Reference(sheet, min_row=1, max_row=sheet.max_row, min_col=2, max_col=2)
    titles = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=1, max_col=1)
    chart1 = BarChart()
    chart1.title = chart_title
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(titles)
    chart1.x_axis.title = x_title
    chart1.y_axis.title = y_title
    #chart1.y_axis.scaling.max = 100
    #chart1.set_categories(titles)
    sheet.add_chart(chart1, 'g2')
    wbook.save('Word_Count.xlsx')

def testing():
    words = mylist(fname)
    stripped_words = chist(words)  # build dictionary
    newlist = extractd(stripped_words)  # build list from dictionary
    sortedlist = sorted(newlist, reverse=True)
    work = exporttoexcel(sortedlist)
    chartme(work, 'Words', 'Word Count', 'Top 5 Popular Words')

testing()

